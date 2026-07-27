"""벌툰(만화/보드게임카페, ㈜아이센스에프앤비) POS "상품별 매출" 파서.

특징:
  - .xlsx (진짜 엑셀). 용량이 커서 한 달 데이터가 여러 파일(날짜 범위)로 쪼개진다.
    예) 2026-05-01~05-14 / 2026-05-15~05-31 → 두 파일을 합쳐야 5월 전체.
  - 레이아웃: 1행 제목(기간 포함), 3행 헤더(번호·대분류·상품코드·상품명·판매수·결제합계
    ·옵션상품코드·옵션상품명·수량·금액), 4행 '합계'(제거), 5행~ 상품.
  - 상품 행 사이에 옵션 행(H~K만 존재, 번호 없음)이 섞여 있다 → 상품 행만 사용.

상품 단위(대분류/상품명/판매수/결제합계) 구조라 PC 분석기(:mod:`app.core.pc_analyzer`)를
그대로 재사용할 수 있도록 :class:`PCParsedFile` 형태로 반환한다.
"""

from __future__ import annotations

import hashlib
import io
import re
import warnings
from dataclasses import dataclass

from openpyxl import load_workbook

from app.core.beltoon_categories import is_menu_category
from app.core.pc_parser import PCCategory, PCParsedFile, PCProduct

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 컬럼(1-based): B..K
_COL = {
    "no": 2, "category": 3, "code": 4, "name": 5,
    "qty": 6, "sales": 7,
    "opt_code": 8, "opt_name": 9, "opt_qty": 10, "opt_amount": 11,
}


class BeltoonParserError(ValueError):
    pass


@dataclass
class _RawProduct:
    code: str
    name: str
    category: str
    qty: float
    sales: float


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    text = str(v).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _find_header_row(ws, max_scan: int = 10) -> int:
    for r in range(1, max_scan + 1):
        vals = {str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else "" for c in range(1, 12)}
        if "상품명" in vals and "판매수" in vals:
            return r
    raise BeltoonParserError("헤더 행(상품명/판매수)을 찾지 못했습니다. 벌툰 파일이 맞는지 확인하세요.")


def _extract_period(ws) -> tuple[str | None, str | None]:
    pat = re.compile(r"(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})")
    for r in range(1, 4):
        for c in range(1, 12):
            text = ws.cell(r, c).value
            if text:
                m = pat.search(str(text))
                if m:
                    return m.group(1), m.group(2)
    return None, None


def _find_total_row(ws, header_row: int) -> tuple[float, float] | None:
    """헤더 다음의 '합계' 행에서 (판매수, 결제합계) 를 읽는다. 없으면 None.

    이 값은 파일이 스스로 밝힌 총계라, 우리가 집계한 상품행 합과 대조해
    행 누락/중복을 즉시 탐지하는 데 쓴다(실측: 정확히 일치).
    """
    for r in range(header_row + 1, min(header_row + 4, ws.max_row + 1)):
        label = _clean(ws.cell(r, _COL["no"]).value)
        if label and "합계" in label:
            return _num(ws.cell(r, _COL["qty"]).value), _num(ws.cell(r, _COL["sales"]).value)
    return None


def _parse_one(source) -> tuple[list[_RawProduct], tuple[str | None, str | None], tuple[float, float] | None]:
    if isinstance(source, bytes):
        wb = load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = load_workbook(source, data_only=True)
    try:
        ws = wb.active
        header = _find_header_row(ws)
        period = _extract_period(ws)
        declared_total = _find_total_row(ws, header)
        out: list[_RawProduct] = []
        for r in range(header + 1, ws.max_row + 1):
            no = ws.cell(r, _COL["no"]).value
            # 상품 행만: 번호가 숫자 (합계행 B='합계', 옵션행 B=None 은 제외)
            if not isinstance(no, (int, float)):
                continue
            name = _clean(ws.cell(r, _COL["name"]).value)
            code = _clean(ws.cell(r, _COL["code"]).value)
            qty = _num(ws.cell(r, _COL["qty"]).value)
            sales = _num(ws.cell(r, _COL["sales"]).value)
            if not name:
                # 이름이 비어도 판매가 있으면 버리지 않는다(매출 누락 방지).
                if not qty and not sales:
                    continue
                name = f"(이름없음) {code}" if code else "(이름없음)"
            out.append(
                _RawProduct(
                    code=code or name,
                    name=name,
                    category=_clean(ws.cell(r, _COL["category"]).value) or "미분류",
                    qty=qty,
                    sales=sales,
                )
            )
        if not out:
            raise BeltoonParserError("상품 데이터를 찾지 못했습니다.")
        return out, period, declared_total
    finally:
        wb.close()


def parse_beltoon_files(sources: list, menu_only: bool = True) -> PCParsedFile:
    """여러 벌툰 파일(같은 달의 분할본)을 합쳐 하나의 :class:`PCParsedFile` 로 반환.

    상품코드 기준으로 판매수·결제합계를 합산한다. 단가는 결제합계/판매수(평균)로 산출.

    Args:
        menu_only: True 면 시간제·이용권 등 비메뉴 분류를 제외하고 메뉴/음료만 분석한다.
                   제외된 항목의 매출/개수/분류수는 결과에 별도로 기록한다.
    """
    if not sources:
        raise BeltoonParserError("파일이 없습니다.")

    merged: dict[str, list] = {}   # code -> [name, category, qty, sales]
    starts: list[str] = []
    ends: list[str] = []
    warns: list[str] = []
    fingerprints: list[str] = []
    seen_fp: set[str] = set()
    ranges: list[tuple[str, str]] = []       # 중복 제외된 파일들의 (시작, 종료)
    declared_qty = declared_sales = 0.0
    have_declared = False
    skipped = 0

    for src in sources:
        # 같은 파일을 두 번 올리면 판매가 그대로 두 배가 되므로 내용 지문으로 차단한다.
        raw_bytes = src if isinstance(src, bytes) else None
        fp = hashlib.sha256(raw_bytes).hexdigest() if raw_bytes is not None else None
        if fp:
            fingerprints.append(fp)
            if fp in seen_fp:
                skipped += 1
                continue
            seen_fp.add(fp)

        raws, (start, end), declared = _parse_one(src)
        if start:
            starts.append(start)
        if end:
            ends.append(end)
        if start and end:
            ranges.append((start, end))
        if declared:
            declared_qty += declared[0]
            declared_sales += declared[1]
            have_declared = True
        for rp in raws:
            acc = merged.setdefault(rp.code, [rp.name, rp.category, 0.0, 0.0])
            acc[2] += rp.qty
            acc[3] += rp.sales

    if skipped:
        warns.append(
            f"⚠️ 같은 파일이 {skipped}개 중복 업로드되어 제외했습니다(그대로 합치면 매출이 "
            f"두 배로 부풀려집니다). 서로 다른 기간의 분할 파일만 올려주세요."
        )

    # 분할 파일의 기간이 겹치면 겹친 구간이 이중 집계된다(내용은 달라도 위험).
    for i in range(len(ranges)):
        for j in range(i + 1, len(ranges)):
            a, b = ranges[i], ranges[j]
            if a[0] <= b[1] and b[0] <= a[1]:
                warns.append(
                    f"⚠️ 업로드한 파일들의 기간이 겹칩니다({a[0]}~{a[1]} / {b[0]}~{b[1]}). "
                    f"겹친 구간이 이중으로 합산됩니다."
                )
                break
        else:
            continue
        break

    # 파일이 스스로 밝힌 '합계'와 우리가 집계한 상품행 합을 대조(행 누락/중복 탐지).
    if have_declared:
        our_qty = sum(v[2] for v in merged.values())
        our_sales = sum(v[3] for v in merged.values())
        if abs(our_sales - declared_sales) > 1.0 or abs(our_qty - declared_qty) > 1.0:
            warns.append(
                f"⚠️ 파일 '합계'행과 집계가 일치하지 않습니다 — 판매수 {our_qty:,.0f} vs "
                f"{declared_qty:,.0f}, 결제합계 {our_sales:,.0f} vs {declared_sales:,.0f}."
            )

    # 메뉴/비메뉴 분리
    excl_sales = excl_qty = 0.0
    excl_cats: set[str] = set()

    products: list[PCProduct] = []
    cat: dict[str, list[float]] = {}  # 메뉴 분류 집계
    for name, category, qty, sales in merged.values():
        if menu_only and not is_menu_category(category):
            excl_sales += sales
            excl_qty += qty
            excl_cats.add(category)
            continue
        products.append(
            PCProduct(name=name, unit_price=(sales / qty) if qty else 0.0, qty=qty, sales=sales)
        )
        c = cat.setdefault(category, [0.0, 0.0])
        c[0] += qty
        c[1] += sales

    categories = [PCCategory(name=n, qty=q, sales=s) for n, (q, s) in cat.items()]

    period = None
    if starts and ends:
        period = f"{min(starts)} ~ {max(ends)}"

    return PCParsedFile(
        products=products,
        categories=categories,
        output_date=period,
        excluded_sales=excl_sales,
        excluded_qty=excl_qty,
        excluded_category_count=len(excl_cats),
        excluded_category_names=sorted(excl_cats),
        warnings=warns,
        fingerprints=fingerprints,
    )
