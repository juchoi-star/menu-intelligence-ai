"""POS "메뉴별 매출 순위 집계" 엑셀 파서.

POS에서 추출되는 엑셀은 대체로 아래 레이아웃을 가지지만, 양식/버전에 따라
컬럼 순서·유무·헤더 라벨 표기가 달라질 수 있어 **컬럼은 고정 위치가 아니라
헤더 이름으로 찾는다**(:data:`_HEADER_ALIASES`, :func:`_build_col_map`).

  - 1~6행   : 리포트 메타(제목, 조회일자, 출력시간 등)
  - 7~8행   : 2단 병합 헤더
  - 9행~     : 데이터 (가맹점 → 메뉴분류 → 메뉴 계층)

데이터 계층 구조
  - 가맹점(가맹점코드/가맹점명)은 그룹의 첫 행에만 표기 → forward-fill.
    가맹점 컬럼 자체가 없는 양식(매장 통합 리포트)도 있는데, 이 경우
    조회범위(scope, 보통 '전체 가맹점')를 단일 가상 매장명으로 사용한다.
  - 메뉴분류는 분류 블록의 첫 행에만 표기 → forward-fill
  - 각 분류 블록은 ``Total`` 행으로, 각 가맹점은 ``소계`` 행으로 끝난다.
  - 소계 뒤에는 비율이 200% 로 찍히는 중복 롤업 행이 하나 더 있다.
  - 파일 끝에는 '총 주문건수 / 총 주문금액 ...' 총계 블록이 있다.

분석 시 Total / 소계 / 총계 / 중복 롤업 행은 모두 제거한다.

깔끔한 단일 판별 규칙:
    유효 메뉴 행 == 메뉴코드가 존재하고, 'Total'이 아니며, 메뉴명이 존재.
이 규칙 하나로 Total·소계·총계·중복행이 전부 걸러진다(실제 3~6월 파일로 검증됨).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# 컬럼 매핑 (1-based, openpyxl 기준). 포맷이 고정이므로 상수로 관리한다.
# ---------------------------------------------------------------------------
COL = {
    "no": 1,           # A  No
    "store_code": 2,   # B  가맹점코드
    "store_name": 3,   # C  가맹점명
    "category": 4,     # D  메뉴분류
    "menu_code": 5,    # E  메뉴코드
    "menu_name": 6,    # F  메뉴명
    "unit_price": 7,   # G  메뉴단가
    "weight": 8,       # H  중량
    "order_count": 9,  # I  주문 건수 (합계)
    "ratio_store": 11,  # K  주문건수 비율(가맹점) — 중복 롤업행 판별용
    "disposal_count": 12,  # L  폐기건수
    "order_amount": 13,    # M  주문금액 (합계)
    "discount_amount": 14, # N  할인금액 (합계)
    "real_sales": 17,      # Q  실매출액 (합계)
    "net_sales": 20,       # T  순매출
    "vat": 21,             # U  부가세
    "cogs": 22,            # V  매출원가 (합계)
    "gross_profit": 25,    # Y  매출이익 (합계)
    "etc_amount": 28,      # AB 기타금액
    "total_sales": 29,     # AC 총판매금액 (합계)
}

# 유효 행 판별에서 제외할 메뉴코드 토큰
_EXCLUDE_CODE_TOKENS = {"total", "소계", "총계", "합계"}

NUMERIC_FIELDS = (
    "unit_price", "weight", "order_count", "disposal_count", "order_amount",
    "discount_amount", "real_sales", "net_sales", "vat", "cogs",
    "gross_profit", "etc_amount", "total_sales",
)


@dataclass(frozen=True)
class MenuRecord:
    """가맹점 × 메뉴 단위의 정규화된 판매 레코드(long format)."""

    store_code: str
    store_name: str
    category: str
    menu_code: str
    menu_name: str
    unit_price: float = 0.0
    weight: float = 0.0
    order_count: float = 0.0
    disposal_count: float = 0.0
    order_amount: float = 0.0
    discount_amount: float = 0.0
    real_sales: float = 0.0
    net_sales: float = 0.0
    vat: float = 0.0
    cogs: float = 0.0
    gross_profit: float = 0.0
    etc_amount: float = 0.0
    total_sales: float = 0.0

    def as_dict(self) -> dict[str, Any]:
        return {
            "store_code": self.store_code,
            "store_name": self.store_name,
            "category": self.category,
            "menu_code": self.menu_code,
            "menu_name": self.menu_name,
            "unit_price": self.unit_price,
            "weight": self.weight,
            "order_count": self.order_count,
            "disposal_count": self.disposal_count,
            "order_amount": self.order_amount,
            "discount_amount": self.discount_amount,
            "real_sales": self.real_sales,
            "net_sales": self.net_sales,
            "vat": self.vat,
            "cogs": self.cogs,
            "gross_profit": self.gross_profit,
            "etc_amount": self.etc_amount,
            "total_sales": self.total_sales,
        }


@dataclass
class Reconciliation:
    """파싱 결과 합계를 파일 끝 POS 총계 블록과 대조한 결과.

    행 누락/중복 집계를 즉시 탐지하기 위한 정확도 안전장치.
    ``ok`` 가 False 면 우리가 집계한 합계가 POS 원본 총계와 다르다는 뜻이다.
    """

    ok: bool
    checks: list[dict] = field(default_factory=list)  # [{label, pos, ours, diff}]

    @property
    def note(self) -> str | None:
        """불일치 항목만 사람이 읽을 수 있는 경고 문자열로."""
        bad = [c for c in self.checks if abs(c["diff"]) > _RECON_TOLERANCE]
        if not bad:
            return None
        parts = [
            f"{c['label']} 집계 {c['ours']:,.0f} vs 원본 {c['pos']:,.0f}(차이 {c['diff']:+,.0f})"
            for c in bad
        ]
        return "⚠️ 원본 POS 총계와 일치하지 않습니다: " + " / ".join(parts)


@dataclass
class ParsedFile:
    """파싱 결과 컨테이너."""

    period_start: date | None
    period_end: date | None
    scope: str | None                       # 예: '전체 가맹점'
    records: list[MenuRecord] = field(default_factory=list)
    reconciliation: Reconciliation | None = None  # POS 총계 대조 결과(정확도 검증)

    @property
    def period_label(self) -> str:
        """대시보드 표기에 쓰는 'YYYY-MM' 라벨(기간 시작 기준)."""
        if self.period_start:
            return f"{self.period_start.year}-{self.period_start.month:02d}"
        return "unknown"


class ParserError(ValueError):
    """엑셀 포맷이 예상과 다를 때 발생."""


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------
def _clean_str(value: Any) -> str | None:
    """셀 값을 trim 한 문자열로. 공백/None 은 None 반환."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_float(value: Any) -> float:
    """숫자 셀을 float 으로 안전 변환. 통화기호·콤마·%·회계식 음수 방어.

    회계 표기에서 음수는 '(1,000)' 처럼 괄호로 감싸므로 이를 -1000 으로 해석한다
    (할인/반품/조정 행에서 등장 가능).
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")  # 회계식 음수
    if negative:
        text = text[1:-1]
    text = text.replace(",", "").replace("₩", "").replace("%", "").strip()
    if not text:
        return 0.0
    try:
        num = float(text)
    except ValueError:
        return 0.0
    return -num if negative else num


def _find_header_row(ws: Worksheet, max_scan: int = 20) -> int:
    """A열이 'No' 인 헤더 행 번호를 찾는다(2단 헤더의 첫 행)."""
    for r in range(1, max_scan + 1):
        if _clean_str(ws.cell(r, COL["no"]).value) == "No":
            return r
    raise ParserError("헤더 행('No')을 찾지 못했습니다. 파일 포맷을 확인하세요.")


def _extract_period(ws: Worksheet, header_row: int) -> tuple[date | None, date | None]:
    """헤더 위쪽 메타 영역에서 '조회일자 : YYYY-MM-DD ~ YYYY-MM-DD' 를 추출."""
    pattern = re.compile(r"(\d{4})-(\d{2})-(\d{2})\s*~\s*(\d{4})-(\d{2})-(\d{2})")
    for r in range(1, header_row):
        for c in range(1, 4):
            text = _clean_str(ws.cell(r, c).value)
            if text and "조회일자" in text:
                m = pattern.search(text)
                if m:
                    y1, m1, d1, y2, m2, d2 = (int(x) for x in m.groups())
                    return date(y1, m1, d1), date(y2, m2, d2)
    return None, None


def _extract_scope(ws: Worksheet, header_row: int) -> str | None:
    """메타 영역에서 조회 범위(예: '전체 가맹점')를 추출."""
    for r in range(1, header_row):
        text = _clean_str(ws.cell(r, COL["no"]).value)
        if text and "가맹점" in text and "조회" not in text:
            return text
    return None


def _is_numlike(v: Any) -> bool:
    """정수/실수, 또는 '12'·'12.0' 같은 숫자 문자열이면 True(HTML 소스는 값이 문자열)."""
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        return s.replace(".", "", 1).isdigit()
    return False


# 총계 대조 허용 오차(원). 부동소수 반올림 정도만 허용하고 그 이상은 불일치로 본다.
_RECON_TOLERANCE = 1.0


def _extract_pos_totals(ws: Worksheet) -> dict[str, float]:
    """파일 끝 총계 블록에서 POS가 직접 찍은 총합을 라벨 기준으로 추출.

    실제 양식:
      - '총 주문건수' 행 → 숫자 [총주문건수, 총주문금액] 순서로 등장
      - '총 순매출액' 행 → 숫자 [총실매출(=주문금액-할인)] 등장
    컬럼 위치는 총계 블록에서 데이터 영역과 다르게 재배치되므로, 컬럼이 아니라
    '행 안의 숫자 값 순서'로 읽는다.
    """
    totals: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        label = _clean_str(ws.cell(r, 1).value)
        if not label:
            continue
        key = re.sub(r"\s+", "", label)
        if "총주문건수" not in key and "총순매출" not in key:
            continue
        nums = [
            _to_float(ws.cell(r, c).value)
            for c in range(2, ws.max_column + 1)
            if _is_numlike(ws.cell(r, c).value)
        ]
        nums = [n for n in nums if n]  # 0 제거(빈 셀·비율 등)
        if "총주문건수" in key and len(nums) >= 2:
            totals.setdefault("order_count", nums[0])
            totals.setdefault("order_amount", nums[1])
        elif "총순매출" in key and nums:
            totals.setdefault("real_sales", nums[0])
    return totals


def _reconcile(records: list[MenuRecord], ws: Worksheet) -> Reconciliation | None:
    """우리가 집계한 합계를 POS 총계 블록과 대조. 총계 블록이 없으면 None."""
    pos = _extract_pos_totals(ws)
    if not pos:
        return None
    ours = {
        "order_count": sum(r.order_count for r in records),
        "order_amount": sum(r.order_amount for r in records),
        "real_sales": sum(r.real_sales for r in records),
    }
    labels = {"order_count": "주문건수", "order_amount": "주문금액", "real_sales": "실매출"}
    checks: list[dict] = []
    for field_key, pos_val in pos.items():
        our_val = ours.get(field_key, 0.0)
        checks.append({
            "label": labels.get(field_key, field_key),
            "pos": pos_val,
            "ours": our_val,
            "diff": our_val - pos_val,
        })
    ok = all(abs(c["diff"]) <= _RECON_TOLERANCE for c in checks)
    return Reconciliation(ok=ok, checks=checks)


def _is_aggregate_row(
    no_value: Any,
    menu_code: str | None,
    category_raw: str | None,
    ratio_store: Any,
) -> bool:
    """집계/요약 행이면 True.

    제외 대상:
      - 총계 블록      : No(A)가 정수 시퀀스가 아님(텍스트/공백)
      - 분류 Total     : 메뉴코드(E) == 'Total'
      - 가맹점 소계    : 메뉴분류(D) == '소계'
      - 중복 롤업 행   : 비율(가맹점)(K) 에 '200%' 포함
    이 방식은 메뉴코드/메뉴명이 비어있는 실제 판매 행(POS 데이터 결손)은 보존한다.
    """
    if not _is_numlike(no_value):
        return True
    if menu_code and menu_code.strip().lower() in _EXCLUDE_CODE_TOKENS:
        return True
    if category_raw and category_raw.strip() == "소계":
        return True
    if "200%" in str(ratio_store or ""):
        return True
    return False


# 헤더 라벨(공백 제거) → 필드 키. POS 양식·버전별로 라벨이 조금씩 달라(예: '실매출액' vs
# '실매출합계') 컬럼 위치가 달라도 헤더로 찾는다. 같은 필드의 라벨 변형은 리스트로 등록.
_HEADER_ALIASES: dict[str, list[str]] = {
    "no": ["No"],
    "store_code": ["가맹점코드"],
    "store_name": ["가맹점명"],
    "category": ["메뉴분류"],
    "menu_code": ["메뉴코드"],
    "menu_name": ["메뉴명"],
    "unit_price": ["메뉴단가"],
    "weight": ["중량"],
    "order_count": ["주문건수"],
    "disposal_count": ["폐기건수"],
    "order_amount": ["주문금액"],
    "discount_amount": ["할인금액"],
    "real_sales": ["실매출액", "실매출합계"],
    "net_sales": ["순매출"],
    "vat": ["부가세"],
    "cogs": ["매출원가"],
    "gross_profit": ["매출이익"],
    "etc_amount": ["기타금액"],
    "total_sales": ["총판매금액"],
}
_HEADER_FIELD: dict[str, str] = {
    label: field for field, labels in _HEADER_ALIASES.items() for label in labels
}


def _build_col_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """헤더 행의 라벨을 읽어 필드→컬럼 인덱스 맵을 만든다(고정 위치가 아니라 이름 기반).

    POS 양식마다 컬럼 순서/유무가 달라도 헤더 텍스트로 정확히 찾아 오정렬을 방지한다.
    """
    cmap: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        if raw is None:
            continue
        label = re.sub(r"\s+", "", str(raw))
        field = _HEADER_FIELD.get(label)
        if field and field not in cmap:
            cmap[field] = c

    if "menu_name" not in cmap or "real_sales" not in cmap:
        raise ParserError(
            "헤더에서 '메뉴명'/'실매출액' 컬럼을 찾지 못했습니다. POS 양식(헤더)을 확인하세요."
        )
    cmap.setdefault("no", 1)
    # 비율(가맹점) 컬럼 = 주문건수(합계) + 2 (합계·비율(분류)·비율(가맹점) 3단 구조). 중복 롤업행 판별용.
    if "order_count" in cmap:
        cmap["ratio_store"] = cmap["order_count"] + 2
    return cmap


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------
def parse_worksheet(ws: Worksheet) -> ParsedFile:
    """워크시트 하나를 파싱하여 정규화된 :class:`ParsedFile` 반환."""
    header_row = _find_header_row(ws)
    data_start = header_row + 2  # 2단 병합 헤더이므로 +2
    cmap = _build_col_map(ws, header_row)
    period_start, period_end = _extract_period(ws, header_row)
    scope = _extract_scope(ws, header_row)

    def cell(r: int, field: str):
        c = cmap.get(field)
        return ws.cell(r, c).value if c else None

    # 일부 POS 양식(예: 매장 통합 리포트)은 가맹점코드/명 컬럼이 아예 없다.
    # 이 경우 조회범위(scope, 보통 '전체 가맹점')를 단일 가상 매장으로 취급한다.
    has_store_col = "store_code" in cmap
    single_store_name = scope or "전체"

    records: list[MenuRecord] = []
    cur_store_code: str | None = None if has_store_col else single_store_name
    cur_store_name: str | None = None if has_store_col else single_store_name
    cur_category: str | None = None
    # 메뉴 forward-fill: 한 메뉴가 여러 행에 걸칠 수 있고(가격 변동 등) 이름·코드는
    # 첫 행에만 있다. 이후 이름·코드 없이 가격만 있는 행은 같은 메뉴의 추가 가격 행이므로
    # 직전 메뉴명/코드를 물려받는다(다음 메뉴명·분류·가맹점이 나오면 리셋).
    cur_menu_code: str | None = None
    cur_menu_name: str | None = None

    for r in range(data_start, ws.max_row + 1):
        no_value = cell(r, "no")
        store_code = _clean_str(cell(r, "store_code"))
        store_name = _clean_str(cell(r, "store_name"))
        category = _clean_str(cell(r, "category"))
        menu_code = _clean_str(cell(r, "menu_code"))
        menu_name = _clean_str(cell(r, "menu_name"))
        ratio_store = cell(r, "ratio_store")

        # forward-fill: 가맹점이 바뀌면 분류·메뉴 컨텍스트를 리셋한다.
        if store_code:
            cur_store_code = store_code
            cur_store_name = store_name or cur_store_name
            cur_category = None
            cur_menu_code = cur_menu_name = None
        # '소계' 는 분류명이 아니라 소계 라벨이므로 forward-fill 대상에서 제외.
        # 새 분류 블록이 시작되면 메뉴 컨텍스트도 리셋(분류 경계를 넘어 물려받지 않음).
        if category and category != "소계":
            cur_category = category
            cur_menu_code = cur_menu_name = None

        if _is_aggregate_row(no_value, menu_code, category, ratio_store):
            continue
        if cur_store_code is None:
            # 헤더 직후 이상 데이터 방어.
            continue

        numeric = {f: _to_float(cell(r, f)) for f in NUMERIC_FIELDS}

        # 완전한 공백 스페이서 행(식별자·판매 모두 없음)은 건너뛴다.
        if (
            not menu_code
            and not menu_name
            and numeric["order_count"] == 0
            and numeric["real_sales"] == 0
        ):
            continue

        # POS 데이터 결손/구조로 메뉴코드·명이 비어도 판매가 있으면 보존.
        #  - 이름이 있으면 새 메뉴 → 그대로 사용하고 forward-fill 기준으로 등록.
        #  - 이름·코드가 둘 다 없으면(가격만) 직전 메뉴의 추가 가격 행이므로 그 메뉴명/코드를
        #    물려받는다(예: 페퍼로니피자감자전 12,900원 다음 줄의 13,900원 인상분 → 페퍼로니로 합산).
        #  - 물려받을 직전 메뉴가 없으면(블록 첫 행이 공백) 분류 기반 '기타 <분류> (<단가>원)'.
        #  - 이름·직전메뉴·분류 모두 없을 때만 "(미상 메뉴)"로 두어 순위에서 제외한다.
        effective_category = cur_category or "미분류"
        if menu_name:
            resolved_name = menu_name
            resolved_code = menu_code or f"NAME-{cur_store_code}-{no_value}"
            cur_menu_name = resolved_name          # 이후 공백 가격행이 물려받을 기준
            cur_menu_code = resolved_code
        elif not menu_code and cur_menu_name:
            # 이름·코드 없는 가격 변동 행 → 직전 메뉴에 합산(forward-fill)
            resolved_name = cur_menu_name
            resolved_code = cur_menu_code
        elif effective_category != "미분류":
            unit_price = int(round(numeric.get("unit_price") or 0))
            if unit_price > 0:
                resolved_name = f"기타 {effective_category} ({unit_price:,}원)"
                resolved_code = menu_code or f"ETC-{effective_category}-{unit_price}"
            else:
                resolved_name = f"기타 {effective_category}"
                resolved_code = menu_code or f"ETC-{effective_category}"
        else:
            resolved_name = "(미상 메뉴)"
            resolved_code = menu_code or f"UNKNOWN-{cur_store_code}-{no_value}"

        records.append(
            MenuRecord(
                store_code=cur_store_code,
                store_name=cur_store_name or cur_store_code,
                category=effective_category,
                menu_code=resolved_code,
                menu_name=resolved_name,
                **numeric,
            )
        )

    if not records:
        raise ParserError("유효한 메뉴 데이터 행을 찾지 못했습니다.")

    return ParsedFile(
        period_start=period_start,
        period_end=period_end,
        scope=scope,
        records=records,
        reconciliation=_reconcile(records, ws),
    )


class _Cell:
    """openpyxl 셀 최소 인터페이스(.value)."""

    __slots__ = ("value",)

    def __init__(self, value: Any) -> None:
        self.value = value


class _GridSheet:
    """행렬(grid) 위에 openpyxl Worksheet 최소 인터페이스를 구현.

    구형 .xls(xlrd)·HTML 표를 openpyxl 과 동일한 방식으로 파싱하기 위한 어댑터.
    """

    def __init__(self, grid: list[list]) -> None:
        self._grid = grid
        self.max_row = len(grid)
        self.max_column = max((len(r) for r in grid), default=0)

    def cell(self, row: int, column: int) -> _Cell:  # 1-based
        r, c = row - 1, column - 1
        if 0 <= r < len(self._grid) and 0 <= c < len(self._grid[r]):
            return _Cell(self._grid[r][c])
        return _Cell(None)


def _html_to_grid(html: str) -> list[list]:
    """HTML 표(.xls/.xlsx 로 위장된)를 행렬로 추출."""
    from html.parser import HTMLParser

    class _T(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.rows: list[list] = []
            self._row: list | None = None
            self._cell: str | None = None

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self._row = []
            elif tag in ("td", "th") and self._row is not None:
                self._cell = ""

        def handle_endtag(self, tag):
            if tag == "tr" and self._row is not None:
                self.rows.append(self._row)
                self._row = None
            elif tag in ("td", "th") and self._cell is not None:
                self._row.append(self._cell.strip())  # type: ignore[union-attr]
                self._cell = None

        def handle_data(self, data):
            if self._cell is not None:
                self._cell += data

    parser = _T()
    parser.feed(html)
    return parser.rows


def _load_worksheet(source: str | bytes):
    """파일 형식(xlsx / 구형 xls / HTML)을 매직바이트로 감지해 (sheet, workbook|None) 반환.

    POS 양식이 진짜 xlsx(zip), 구형 .xls(OLE/BIFF), 또는 HTML 표로 내보내질 수 있어
    확장자가 아니라 실제 내용으로 판별한다.
    """
    import io

    data = source if isinstance(source, bytes) else open(source, "rb").read()
    head = data[:8]

    if head[:4] == b"PK\x03\x04":  # .xlsx (zip)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        return wb.active, wb

    if head[:4] == b"\xd0\xcf\x11\xe0":  # 구형 .xls (OLE/BIFF)
        import xlrd

        book = xlrd.open_workbook(file_contents=data)
        sh = book.sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return _GridSheet(grid), None

    text = data.decode("utf-8", errors="replace")
    if "<table" in text[:5000].lower() or "<html" in text[:2000].lower():  # HTML 표
        return _GridSheet(_html_to_grid(text)), None

    raise ParserError("지원하지 않는 파일 형식입니다. POS에서 .xlsx 또는 .xls 로 내보내 주세요.")


def parse_excel(source: str | bytes) -> ParsedFile:
    """POS 파일(경로 또는 bytes)을 형식 자동 감지하여 파싱한다.

    Args:
        source: 파일 경로 문자열 또는 업로드된 파일의 raw bytes.
    """
    ws, wb = _load_worksheet(source)
    try:
        if ws is None:
            raise ParserError("활성 시트를 찾을 수 없습니다.")
        return parse_worksheet(ws)
    finally:
        if wb is not None:
            wb.close()
