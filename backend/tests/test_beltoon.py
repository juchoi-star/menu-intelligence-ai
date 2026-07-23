"""벌툰 파서·병합 테스트 (합성 xlsx)."""

from __future__ import annotations

import io

import openpyxl

from app.core.beltoon_parser import parse_beltoon_files


def _make_xlsx(period: str, rows: list[tuple], total_qty: float, total_sales: float) -> bytes:
    """벌툰 포맷 축약 재현: 1행 제목, 3행 헤더, 4행 합계, 5행~ 상품(+옵션행 1개)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 2, f"㈜테스트 외 211의 {period} 상품별 매출 세트포함")
    hdr = ["번호", "대분류", "상품코드", "상품명", "판매수", "결제 합계",
           "옵션상품코드", "옵션상품명", "수량", "금액"]
    for i, h in enumerate(hdr):
        ws.cell(3, 2 + i, h)
    ws.cell(4, 2, "합계"); ws.cell(4, 6, total_qty); ws.cell(4, 7, total_sales)
    r = 5
    for no, (cat, code, name, qty, sales) in enumerate(rows, 1):
        ws.cell(r, 2, no); ws.cell(r, 3, cat); ws.cell(r, 4, code)
        ws.cell(r, 5, name); ws.cell(r, 6, qty); ws.cell(r, 7, sales)
        r += 1
        # 옵션 행(번호 없음, H~K만) — 무시되어야 함
        ws.cell(r, 8, "OPT1"); ws.cell(r, 9, "옵션"); ws.cell(r, 10, 5); ws.cell(r, 11, 0)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_beltoon_merge_sums_by_code():
    # 같은 달의 두 분할본(전반/후반). 같은 상품코드 합산. (메뉴 분류만 사용)
    a = _make_xlsx("2026-05-01 ~ 2026-05-14",
                   [("식사", "A1", "치킨마요덮밥", 100, 500000),
                    ("음료", "B1", "콜라", 50, 100000)], 150, 600000)
    b = _make_xlsx("2026-05-15 ~ 2026-05-31",
                   [("식사", "A1", "치킨마요덮밥", 120, 600000),
                    ("음료", "B1", "콜라", 30, 60000)], 150, 660000)
    pf = parse_beltoon_files([a, b])

    assert pf.output_date == "2026-05-01 ~ 2026-05-31"
    assert len(pf.products) == 2  # 옵션행 제외, 합산
    two = next(p for p in pf.products if p.name == "치킨마요덮밥")
    assert two.qty == 220
    assert two.sales == 1100000
    assert two.unit_price == 5000
    assert sum(p.sales for p in pf.products) == 1260000
    assert {c.name for c in pf.categories} == {"식사", "음료"}


def test_beltoon_excludes_non_menu_categories():
    """시간제·요금·진동벨 등 비메뉴 분류는 제외되고 별도 집계된다."""
    a = _make_xlsx("2026-05-01 ~ 2026-05-31",
                   [("시간제", "T1", "2시간", 1000, 5000000),      # 제외
                    ("성인시간제", "T2", "성인2시간", 100, 800000),  # 제외
                    ("진동벨", "P1", "진동벨1", 50, 10000),          # 제외
                    ("식사", "F1", "제육덮밥", 200, 1400000),        # 유지
                    ("음료", "D1", "콜라", 100, 200000)], 1450, 7410000)  # 유지
    pf = parse_beltoon_files([a])

    names = {p.name for p in pf.products}
    assert names == {"제육덮밥", "콜라"}                 # 메뉴만
    assert sum(p.sales for p in pf.products) == 1600000   # 1,400,000 + 200,000
    assert pf.excluded_sales == 5810000                   # 시간제+성인+진동벨
    assert pf.excluded_category_count == 3
    assert "시간제" in pf.excluded_category_names

    # menu_only=False 면 전체 유지
    pf_all = parse_beltoon_files([a], menu_only=False)
    assert len(pf_all.products) == 5
    assert pf_all.excluded_sales == 0
