"""파서 단위 테스트.

핵심 검증:
  - 집계행(Total/소계/총계/중복 롤업) 제거
  - forward-fill(가맹점/분류)
  - 메뉴코드/명이 비어도 판매가 있으면 보존
"""

from __future__ import annotations

import openpyxl
import pytest

from app.core.parser import ParserError, parse_worksheet


def _make_sheet():
    """실제 POS 포맷을 축약 재현한 워크시트를 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 메타
    ws.cell(1, 1, "메뉴별 매출 순위 집계")
    ws.cell(3, 1, "전체 가맹점")
    ws.cell(4, 1, "조회일자 : 2026-05-01 ~ 2026-05-31   ")
    # 헤더(7행) — 헤더 이름 기반 컬럼 매핑을 위해 라벨을 채운다.
    for c, label in {
        1: "  No  ", 2: "  가맹점코드  ", 3: "  가맹점명  ", 4: "  메뉴분류  ",
        5: "  메뉴코드  ", 6: "  메뉴명  ", 7: "  메뉴단가  ", 9: "  주문 건수  ",
        13: "  주문금액  ", 14: "  할인금액  ", 17: "  실매출액  ", 25: "  매출이익  ",
    }.items():
        ws.cell(7, c, label)
    # 데이터(9행~)  컬럼: A No, B 코드, C 명, D 분류, E 메뉴코드, F 메뉴명,
    #                G 단가, ... I 주문건수(9), K 비율가맹점(11), M 주문금액(13),
    #                N 할인(14), Q 실매출(17), Y 매출이익(25)
    def row(r, no, code, name, cat, mcode, mname, price, oc, ratio_store, oa, disc, real, profit):
        ws.cell(r, 1, no); ws.cell(r, 2, code); ws.cell(r, 3, name); ws.cell(r, 4, cat)
        ws.cell(r, 5, mcode); ws.cell(r, 6, mname); ws.cell(r, 7, price)
        ws.cell(r, 9, oc); ws.cell(r, 11, ratio_store); ws.cell(r, 13, oa)
        ws.cell(r, 14, disc); ws.cell(r, 17, real); ws.cell(r, 25, profit)

    # 가맹점 A / 분류 '전'
    row(9,  1, "F001", "가맹점A", "전", "1001", "김치전", 8000, 10, "5%", 80000, 0, 80000, 72000)
    row(10, 2, None, None, None, "1002", "해물파전", 9000, 5, "3%", 45000, 5000, 40000, 36000)
    # 분류 Total (제거 대상)
    ws.cell(11, 1, 3); ws.cell(11, 5, "Total"); ws.cell(11, 9, 15); ws.cell(11, 17, 120000)
    # 이름 없는 실제 판매행 (보존 대상)
    row(12, 4, None, None, None, None, None, 5000, 2, "1%", 10000, 0, 10000, 9000)
    # 가맹점 소계 (제거 대상)
    ws.cell(13, 1, 5); ws.cell(13, 4, "소계"); ws.cell(13, 9, 17); ws.cell(13, 17, 130000)
    # 중복 롤업 (제거 대상, 비율 200%)
    ws.cell(14, 1, 6); ws.cell(14, 11, "200%"); ws.cell(14, 9, 17); ws.cell(14, 17, 130000)
    # 가맹점 B / 분류 '탕&식사'
    row(15, 7, "F002", "가맹점B", "탕&식사", "2001", "김치찌개", 9000, 20, "10%", 180000, 0, 180000, 162000)
    # 총계 블록 (A가 텍스트, 제거 대상)
    ws.cell(16, 1, "총 주문건수 "); ws.cell(16, 29, 40)
    return ws


def test_parse_removes_aggregate_rows_and_forward_fills():
    ws = _make_sheet()
    parsed = parse_worksheet(ws)

    # 유효 행: 김치전, 해물파전, (미상 메뉴), 김치찌개 = 4개
    assert len(parsed.records) == 4

    names = [r.menu_name for r in parsed.records]
    assert "김치전" in names
    assert "(미상 메뉴)" in names          # 이름 없는 판매행 보존
    assert "김치찌개" in names
    assert not any(n in ("Total", "소계") for n in names)

    # forward-fill 검증
    haemul = next(r for r in parsed.records if r.menu_name == "해물파전")
    assert haemul.store_code == "F001"
    assert haemul.store_name == "가맹점A"
    assert haemul.category == "전"

    # 가맹점 전환 시 분류 컨텍스트 리셋
    jjigae = next(r for r in parsed.records if r.menu_name == "김치찌개")
    assert jjigae.store_code == "F002"
    assert jjigae.category == "탕&식사"

    # 메타
    assert parsed.period_label == "2026-05"
    assert parsed.scope == "전체 가맹점"


def test_totals_reconcile():
    ws = _make_sheet()
    parsed = parse_worksheet(ws)
    total_real = sum(r.real_sales for r in parsed.records)
    # 80000 + 40000 + 10000(미상) + 180000 = 310000
    assert total_real == 310000
    total_orders = sum(r.order_count for r in parsed.records)
    assert total_orders == 37  # 10 + 5 + 2 + 20


def test_empty_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(7, 1, "  No  ")
    with pytest.raises(ParserError):
        parse_worksheet(ws)
